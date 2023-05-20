import openpyxl
import mysql.connector
# ワークシート読み込み
wb = openpyxl.load_workbook('music_chart2022.xlsx')

# セル用のリスト
values = [[] for i in range(10)]

# シート読み込み&リスト挿入
ws = wb.worksheets[0]
for row in range(2, 12):
    # rank挿入
    nowcell = int(ws.cell(row, 1).value)
    values[row - 2].append(nowcell)
    # song挿入
    nowcell = ws.cell(row, 2).value
    values[row - 2].append(nowcell)
print(values)

# SQLへのリスト挿入
cnx = None

try:
    cnx = mysql.connector.connect(
        user='name',  # ユーザー名
        password='password',  # パスワード
        host='localhost',  # ホスト名(IPアドレス）
        database='music'  # データベース名
    )

    cursor = cnx.cursor()

    sql = ('''
    INSERT INTO `global` 
        (`rank`, `song`)
    VALUES 
        (%s, %s)
    ''')

    cursor.executemany(sql, values)
    cnx.commit()

    print(f"{cursor.rowcount} records inserted.")

    cursor.close()

except Exception as e:
    print(f"Error Occurred: {e}")

finally:
    if cnx is not None and cnx.is_connected():
        cnx.close()